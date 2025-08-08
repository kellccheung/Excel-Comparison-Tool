#!/usr/bin/env python3
"""
Test script for Excel Compare Tool

This script creates sample Excel files and tests the comparison functionality.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
import os
from comparison_engine import ComparisonEngine
from report_generator import ReportGenerator


def create_sample_excel_files():
    """Create sample Excel files for testing."""
    
    # Create first sample file
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "Sheet1"
    
    # Add some data
    data1 = [
        ['Name', 'Age', 'City', 'Salary'],
        ['John', 25, 'New York', 50000],
        ['Jane', 30, 'Los Angeles', 60000],
        ['Bob', 35, 'Chicago', 55000],
        ['Alice', 28, 'Boston', 52000]
    ]
    
    for row_idx, row_data in enumerate(data1, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add a formula
    ws1['E1'] = 'Total'
    ws1['E2'] = '=SUM(D2:D5)'
    
    # Create second sheet
    ws2 = wb1.create_sheet("Sheet2")
    ws2['A1'] = 'Product'
    ws2['B1'] = 'Price'
    ws2['A2'] = 'Laptop'
    ws2['B2'] = 1200
    ws2['A3'] = 'Mouse'
    ws2['B3'] = 25
    
    wb1.save('sample_file1.xlsx')
    print("Created sample_file1.xlsx")
    
    # Create second sample file (with differences)
    wb2 = openpyxl.Workbook()
    ws1_2 = wb2.active
    ws1_2.title = "Sheet1"
    
    # Add similar data with some differences
    data2 = [
        ['Name', 'Age', 'City', 'Salary'],
        ['John', 25, 'New York', 50000],
        ['Jane', 30, 'Los Angeles', 65000],  # Different salary
        ['Bob', 35, 'Chicago', 55000],
        ['Alice', 28, 'Boston', 52000],
        ['Charlie', 32, 'Miami', 58000]  # Extra row
    ]
    
    for row_idx, row_data in enumerate(data2, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1_2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add a different formula
    ws1_2['E1'] = 'Total'
    ws1_2['E2'] = '=SUM(D2:D6)'  # Different range
    
    # Create different second sheet
    ws2_2 = wb2.create_sheet("Sheet2")
    ws2_2['A1'] = 'Product'
    ws2_2['B1'] = 'Price'
    ws2_2['A2'] = 'Laptop'
    ws2_2['B2'] = 1200
    ws2_2['A3'] = 'Mouse'
    ws2_2['B3'] = 30  # Different price
    ws2_2['A4'] = 'Keyboard'
    ws2_2['B4'] = 80  # Extra product
    
    # Add a third sheet (only in second file)
    ws3_2 = wb2.create_sheet("Sheet3")
    ws3_2['A1'] = 'Month'
    ws3_2['B1'] = 'Sales'
    ws3_2['A2'] = 'January'
    ws3_2['B2'] = 1000
    ws3_2['A3'] = 'February'
    ws3_2['B3'] = 1200
    
    wb2.save('sample_file2.xlsx')
    print("Created sample_file2.xlsx")


def test_comparison():
    """Test the comparison functionality."""
    print("\n" + "="*50)
    print("TESTING EXCEL COMPARISON")
    print("="*50)
    
    # Create sample files
    create_sample_excel_files()
    
    # Initialize comparison engine
    engine = ComparisonEngine()
    
    # Load files
    print("\nLoading files...")
    if not engine.load_files('sample_file1.xlsx', 'sample_file2.xlsx'):
        print("Failed to load files!")
        return
    
    # Perform comparison
    print("Performing comparison...")
    results = engine.compare_all()
    
    if not results:
        print("Comparison failed!")
        return
    
    # Display results
    print("\nCOMPARISON RESULTS:")
    print("-" * 30)
    
    summary = results.get('summary', {})
    print(f"Files Identical: {summary.get('files_identical', False)}")
    print(f"Total Differences: {summary.get('total_differences', 0)}")
    
    diff_by_type = summary.get('differences_by_type', {})
    for diff_type, count in diff_by_type.items():
        print(f"{diff_type.title()}: {count}")
    
    # Test report generation
    print("\n" + "="*50)
    print("TESTING REPORT GENERATION")
    print("="*50)
    
    export_data = engine.export_comparison_data()
    report_gen = ReportGenerator(export_data)
    
    # Generate text report
    if report_gen.generate_text_report('test_report.txt'):
        print("✓ Text report generated successfully")
    else:
        print("✗ Failed to generate text report")
    
    # Generate Excel report
    if report_gen.generate_excel_report('test_report.xlsx'):
        print("✓ Excel report generated successfully")
    else:
        print("✗ Failed to generate Excel report")
    
    # Clean up
    engine.close()
    
    print("\n" + "="*50)
    print("TEST COMPLETED")
    print("="*50)
    print("Generated files:")
    print("- sample_file1.xlsx")
    print("- sample_file2.xlsx")
    print("- test_report.txt")
    print("- test_report.xlsx")
    print("\nYou can now run the GUI application with: python excel_compare.py")


if __name__ == "__main__":
    test_comparison()
