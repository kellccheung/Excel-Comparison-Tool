import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, List, Tuple, Any, Optional
import re
import os


class ExcelParser:
    """Parser for Excel files with support for formulas and VBA code extraction."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.vba_code = {}
        self.formulas = {}
        self.sheet_data = {}
        
    def load_workbook(self) -> bool:
        """Load the Excel workbook and extract all data."""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=False)
            self._extract_sheet_data()
            self._extract_formulas()
            self._extract_vba_code()
            return True
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
    
    def _extract_sheet_data(self):
        """Extract data from all sheets."""
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            data = []
            
            # Get the used range
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            
            # Convert to DataFrame for easier handling
            df = pd.DataFrame(data)
            self.sheet_data[sheet_name] = df
    
    def _extract_formulas(self):
        """Extract formulas from all sheets."""
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            formulas = {}
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and cell.data_type == 'f':  # Formula
                        cell_address = f"{get_column_letter(cell.column)}{cell.row}"
                        formulas[cell_address] = {
                            'formula': cell.value,
                            'value': cell.value if hasattr(cell, 'value') else None,
                            'sheet': sheet_name
                        }
            
            if formulas:
                self.formulas[sheet_name] = formulas
    
    def _extract_vba_code(self):
        """Extract VBA code from the Excel file."""
        try:
            with zipfile.ZipFile(self.file_path, 'r') as zip_file:
                # Look for VBA project file
                vba_files = [f for f in zip_file.namelist() if 'vbaProject.bin' in f]
                
                if vba_files:
                    # Extract VBA project
                    vba_data = zip_file.read(vba_files[0])
                    self.vba_code['vba_project'] = vba_data
                
                # Look for VBA modules in the project
                for file_name in zip_file.namelist():
                    if 'xl/vba/' in file_name and file_name.endswith('.bin'):
                        module_name = os.path.basename(file_name).replace('.bin', '')
                        try:
                            module_data = zip_file.read(file_name)
                            self.vba_code[module_name] = module_data
                        except:
                            continue
                            
        except Exception as e:
            print(f"Error extracting VBA code: {e}")
    
    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names."""
        return self.workbook.sheetnames if self.workbook else []
    
    def get_sheet_data(self, sheet_name: str) -> pd.DataFrame:
        """Get data from a specific sheet."""
        return self.sheet_data.get(sheet_name, pd.DataFrame())
    
    def get_formulas(self, sheet_name: str = None) -> Dict:
        """Get formulas from a specific sheet or all sheets."""
        if sheet_name:
            return self.formulas.get(sheet_name, {})
        return self.formulas
    
    def get_vba_code(self) -> Dict:
        """Get all VBA code."""
        return self.vba_code
    
    def get_cell_info(self, sheet_name: str, cell_address: str) -> Dict:
        """Get detailed information about a specific cell."""
        if sheet_name not in self.workbook.sheetnames:
            return {}
        
        sheet = self.workbook[sheet_name]
        try:
            cell = sheet[cell_address]
            info = {
                'value': cell.value,
                'data_type': cell.data_type,
                'formula': cell.value if cell.data_type == 'f' else None,
                'format': str(cell.number_format) if hasattr(cell, 'number_format') else None,
                'font': {
                    'bold': cell.font.bold if hasattr(cell.font, 'bold') else None,
                    'italic': cell.font.italic if hasattr(cell.font, 'italic') else None,
                    'size': cell.font.size if hasattr(cell.font, 'size') else None,
                    'color': cell.font.color.rgb if hasattr(cell.font, 'color') and hasattr(cell.font.color, 'rgb') else None
                },
                'fill': {
                    'type': cell.fill.fill_type if hasattr(cell.fill, 'fill_type') else None,
                    'color': cell.fill.start_color.rgb if hasattr(cell.fill, 'start_color') and hasattr(cell.fill.start_color, 'rgb') else None
                }
            }
            return info
        except:
            return {}
    
    def get_workbook_properties(self) -> Dict:
        """Get workbook properties."""
        if not self.workbook:
            return {}
        
        return {
            'title': self.workbook.properties.title,
            'subject': self.workbook.properties.subject,
            'creator': self.workbook.properties.creator,
            'created': self.workbook.properties.created,
            'modified': self.workbook.properties.modified,
            'last_modified_by': self.workbook.properties.lastModifiedBy,
            'sheets_count': len(self.workbook.sheetnames)
        }
    
    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()
